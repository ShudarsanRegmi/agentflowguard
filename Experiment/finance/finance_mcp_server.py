import os
import sqlite3
import json
from mcp.server.fastmcp import FastMCP
from openpyxl import load_workbook

mcp = FastMCP("Finance Agent MCP")
DB_PATH = "/home/aparichit/Projects/AgentFlowGuard/Experiment/finance/finance.db"
FINANCE_DIR = "/home/aparichit/Projects/AgentFlowGuard/Experiment/finance"

@mcp.tool()
def query_finance_db(sql_query: str) -> str:
    """Execute SQL queries on the Finance database.
    Supported tables: Employees, Payroll, Expenses, Transactions.
    """
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        c.execute(sql_query)
        rows = c.fetchall()
        conn.close()
        
        if not rows:
            return "Query executed successfully, no rows returned."
        
        res = [dict(row) for row in rows]
        return json.dumps(res, indent=2)
    except Exception as e:
        return f"Error executing query: {e}"

@mcp.tool()
def read_salary_spreadsheet() -> str:
    """Parse and return structured data from the salary.xlsx spreadsheet."""
    xlsx_path = os.path.join(FINANCE_DIR, "salary.xlsx")
    try:
        wb = load_workbook(xlsx_path, read_only=True)
        sheet = wb.active
        data = []
        for row in sheet.iter_rows(values_only=True):
            if row:
                data.append(list(row))
        
        # Format as simple table/rows string
        res = "\n".join([" | ".join([str(val) if val is not None else "" for val in r]) for r in data])
        return res
    except Exception as e:
        return f"Error reading salary.xlsx: {e}"

@mcp.tool()
def read_tax_records_pdf() -> str:
    """Read the tax records PDF details containing gross profits, FED id, and deductions."""
    pdf_path = os.path.join(FINANCE_DIR, "tax_records.pdf")
    try:
        import pypdf
        reader = pypdf.PdfReader(pdf_path)
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"
        return text
    except Exception as e:
        return f"Error reading tax_records.pdf: {e}"

@mcp.tool()
def list_invoices() -> str:
    """List all available invoices in the invoices directory."""
    invoices_dir = os.path.join(FINANCE_DIR, "invoices")
    try:
        files = os.listdir(invoices_dir)
        return json.dumps(files, indent=2)
    except Exception as e:
        return f"Error listing invoices: {e}"

@mcp.tool()
def read_invoice(filename: str) -> str:
    """Read content of a specific invoice file by filename."""
    invoice_path = os.path.join(FINANCE_DIR, "invoices", filename)
    # Basic path safety check
    if not os.path.abspath(invoice_path).startswith(os.path.abspath(FINANCE_DIR)):
        return "Access denied: outside directory boundary."
    try:
        import pypdf
        reader = pypdf.PdfReader(invoice_path)
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"
        return text
    except Exception as e:
        return f"Error reading invoice file {filename}: {e}"

@mcp.tool()
def list_receipts() -> str:
    """List all available receipts in the receipts directory."""
    receipts_dir = os.path.join(FINANCE_DIR, "receipts")
    try:
        files = os.listdir(receipts_dir)
        return json.dumps(files, indent=2)
    except Exception as e:
        return f"Error listing receipts: {e}"

@mcp.tool()
def read_receipt(filename: str) -> str:
    """Read content of a specific receipt file by filename."""
    receipt_path = os.path.join(FINANCE_DIR, "receipts", filename)
    if not os.path.abspath(receipt_path).startswith(os.path.abspath(FINANCE_DIR)):
        return "Access denied: outside directory boundary."
    try:
        import pypdf
        reader = pypdf.PdfReader(receipt_path)
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"
        return text
    except Exception as e:
        return f"Error reading receipt file {filename}: {e}"

if __name__ == "__main__":
    mcp.run()
